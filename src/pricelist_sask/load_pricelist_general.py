#!/usr/bin/env python3
import argparse
from pathlib import Path
import re
from typing import Any, Dict, List, Optional

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# Database connection defaults from SLGApricelist.md
DB_CONFIG = {
    'host': 'localhost',
    'database': 'bottlebridge',
    'user': 'olena',
    'password': 'Barnet359?',
}

TARGET_TABLE = 'public.pricelist_general'

HEADER_TO_FIELD = {
    'Item': 'cspcid',
    'UPC': 'pid',
    'Item Description': 'description',
    'Size (mL)': 'size_id',
    'Refund Dep': 'deposit',
    'Wholesale Base Price': 'wholesale',
    'Product Hierarchy': 'product_hierarchy',
    'Manufacturer Name': 'supplier',
    'Sweetness': 'sweetness',
    'Alcohol Content (%)': 'alcohol',
    'Units/Case': 'multiple',
}

PROVINCE = 'SASK'

INSERT_COLUMNS = [
    'price_date',
    'cspcid',
    'pid',
    'description',
    'size_id',
    'deposit',
    'wholesale',
    'category',
    'group',
    'supplier',
    'sweetness',
    'alcohol',
    'multiple',
    'province',
]


def parse_price_date_from_filename(path: Path) -> str:
    match = re.search(r'(\d{8})', path.name)
    if not match:
        raise ValueError(f'Unable to parse price_date from filename: {path.name}')
    return match.group(1)


def parse_product_hierarchy(value: Optional[str]) -> Dict[str, str]:
    if value is None:
        return {'group': '', 'category': ''}

    value_str = str(value).strip()
    if not value_str:
        return {'group': '', 'category': ''}

    parts = [part.strip() for part in value_str.split(' - ', 1)]
    group = parts[0]
    category = parts[1] if len(parts) > 1 else ''
    return {'group': group, 'category': category}


def build_row(header_map: Dict[str, int], row_values: List[Any], price_date: str) -> Optional[List[Any]]:
    raw: Dict[str, Any] = {}
    for header, idx in header_map.items():
        raw[header] = row_values[idx] if idx < len(row_values) else None

    if raw.get('UPC') is None and raw.get('Item') is None and raw.get('Item Description') is None:
        return None

    hierarchy = parse_product_hierarchy(raw.get('Product Hierarchy'))

    return [
        price_date,
        raw.get('Item'),
        raw.get('UPC'),
        raw.get('Item Description'),
        raw.get('Size (mL)'),
        raw.get('Refund Dep'),
        raw.get('Wholesale Base Price'),
        hierarchy['category'],
        hierarchy['group'],
        raw.get('Manufacturer Name'),
        raw.get('Sweetness'),
        raw.get('Alcohol Content (%)'),
        raw.get('Units/Case'),
        PROVINCE,
    ]


def load_workbook_rows(path: Path) -> List[List[Any]]:
    workbook = openpyxl.load_workbook(filename=path, data_only=True)
    sheet = workbook.active

    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map: Dict[str, int] = {}
    for idx, header in enumerate(header_row):
        if header in HEADER_TO_FIELD:
            header_map[header] = idx

    missing_headers = [key for key in HEADER_TO_FIELD if key not in header_map]
    if missing_headers:
        raise ValueError(f'Missing expected headers in workbook: {missing_headers}')

    price_date = parse_price_date_from_filename(path)
    rows: List[List[Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        parsed = build_row(header_map, list(row), price_date)
        if parsed is not None:
            rows.append(parsed)

    return rows


def load_into_db(rows: List[List[Any]]) -> None:
    if not rows:
        print('No rows to insert.')
        return

    col_list = ', '.join('"' + col + '"' if col == 'group' else col for col in INSERT_COLUMNS)

    # Update columns are everything except the conflict key (cspcid, pid)
    update_cols = [c for c in INSERT_COLUMNS if c not in ('cspcid', 'pid')]

    def update_expr(col: str) -> str:
        quoted = f'"{col}"' if col == 'group' else col
        t = TARGET_TABLE
        # Skip update if new value is NULL or empty string
        return (
            f"{quoted} = CASE WHEN EXCLUDED.{quoted} IS NOT NULL AND EXCLUDED.{quoted}::text != '' "
            f"THEN EXCLUDED.{quoted} ELSE {TARGET_TABLE}.{quoted} END"
        )

    update_set = ', '.join(update_expr(c) for c in update_cols)

    upsert_query = f"""
        INSERT INTO {TARGET_TABLE} ({col_list})
        VALUES %s
        ON CONFLICT (cspcid, pid) DO UPDATE SET {update_set}
    """

    with psycopg2.connect(**DB_CONFIG) as conn:
        with conn.cursor() as cursor:
            execute_values(cursor, upsert_query, rows)
        conn.commit()

    print(f'Upserted {len(rows)} rows into {TARGET_TABLE}.')


def main() -> int:
    parser = argparse.ArgumentParser(description='Load SLGA pricelist data into pricelist_general.')
    parser.add_argument('files', nargs='+', type=Path, help='Path(s) to .xls/.xlsx pricelist file(s)')
    args = parser.parse_args()

    for input_file in args.files:
        if not input_file.exists():
            print(f'Input file not found: {input_file}')
            return 1

        rows = load_workbook_rows(input_file)
        print(f'Parsed {len(rows)} rows from {input_file.name}.')
        load_into_db(rows)

    return 0


if __name__ == '__main__':
    raise SystemExit(main())

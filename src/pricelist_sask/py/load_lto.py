#!/usr/bin/env python3
import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DB_CONFIG = {
    'host': 'localhost',
    'database': 'bottlebridge',
    'user': 'olena',
    'password': 'Barnet359?',
}

TARGET_TABLE = 'public.LTO_sk'

HEADER_TO_FIELD = {
    'WPP Start Date': 'date_from',
    'WPP End Date': 'date_to',
    'SLGA Item No.': 'cspcid',
    'Wholesale Base Price': 'wholesale',
    'WPP Savings': 'LTO',
}

INSERT_COLUMNS = ['cspcid', 'date_from', 'date_to', 'lto', 'wholesale']
CONFLICT_COLUMNS = ['cspcid', 'date_from', 'date_to']


def build_row(header_map: Dict[str, int], row_values: List[Any]) -> Optional[List[Any]]:
    raw: Dict[str, Any] = {}
    for header, idx in header_map.items():
        raw[header] = row_values[idx] if idx < len(row_values) else None

    if all(raw.get(h) is None for h in HEADER_TO_FIELD):
        return None

    return [
        raw.get('SLGA Item No.'),
        raw.get('WPP Start Date'),
        raw.get('WPP End Date'),
        raw.get('WPP Savings'),
        raw.get('Wholesale Base Price'),
    ]


def load_workbook_rows(path: Path) -> List[List[Any]]:
    workbook = openpyxl.load_workbook(filename=path, data_only=True)
    sheet = workbook.active

    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map: Dict[str, int] = {}
    for idx, header in enumerate(header_row):
        if header in HEADER_TO_FIELD:
            header_map[header] = idx

    missing_headers = [key for key in HEADER_TO_FIELD if key not in header_map]
    if missing_headers:
        raise ValueError(f'Missing expected headers in workbook: {missing_headers}')

    rows: List[List[Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        parsed = build_row(header_map, list(row))
        if parsed is not None:
            rows.append(parsed)

    return rows


def load_into_db(rows: List[List[Any]]) -> None:
    if not rows:
        print('No rows to process.')
        return

    with psycopg2.connect(**DB_CONFIG) as conn:
        with conn.cursor() as cursor:
            # Stage into a temp table
            cursor.execute("""
                CREATE TEMP TABLE tmp_lto (
                    cspcid    varchar(32),
                    date_from date,
                    date_to   date,
                    lto       numeric(12,5),
                    wholesale numeric(12,5)
                ) ON COMMIT DROP
            """)
            execute_values(
                cursor,
                "INSERT INTO tmp_lto (cspcid, date_from, date_to, lto, wholesale) VALUES %s",
                rows,
            )

            # Update existing rows; only overwrite lto/wholesale when incoming value is not null
            cursor.execute(f"""
                UPDATE {TARGET_TABLE} t
                SET lto = CASE
                        WHEN s.lto IS NOT NULL THEN s.lto
                        ELSE t.lto
                      END,
                    wholesale = CASE
                        WHEN s.wholesale IS NOT NULL THEN s.wholesale
                        ELSE t.wholesale
                      END
                FROM tmp_lto s
                WHERE t.cspcid    = s.cspcid
                  AND t.date_from = s.date_from
                  AND t.date_to   = s.date_to
            """)
            updated = cursor.rowcount

            # Insert rows that don't exist yet
            cursor.execute(f"""
                INSERT INTO {TARGET_TABLE} (cspcid, date_from, date_to, lto, wholesale)
                SELECT s.cspcid, s.date_from, s.date_to, s.lto, s.wholesale
                FROM tmp_lto s
                LEFT JOIN {TARGET_TABLE} t
                       ON t.cspcid    = s.cspcid
                      AND t.date_from = s.date_from
                      AND t.date_to   = s.date_to
                WHERE t.id IS NULL
            """)
            inserted = cursor.rowcount

        conn.commit()

    print(f'Done: {inserted} inserted, {updated} updated in {TARGET_TABLE}.')


def main() -> int:
    parser = argparse.ArgumentParser(description='Load SLGA LTO data into lto_sk.')
    parser.add_argument('files', nargs='+', type=Path, help='Path(s) to .xls/.xlsx LTO file(s)')
    args = parser.parse_args()

    for input_file in args.files:
        if not input_file.exists():
            print(f'Input file not found: {input_file}')
            return 1

        rows = load_workbook_rows(input_file)
        print(f'Parsed {len(rows)} rows from {input_file.name}.')
        load_into_db(rows)

    return 0


if __name__ == '__main__':
    raise SystemExit(main())
